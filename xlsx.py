from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

file_path = 'xlsx/test1.xlsx'
new_file_path = 'xlsx/test2.xlsx'
img_path = 'images/title.png'
wb = load_workbook(file_path)
ws = wb['Sheet1']
ws.print_title_rows = "1:1"

#从数据源获取的其他信息
data={
    'C4' : '002LCNB250350001',
    'D5' : 'EC-202501-STEG-LC-COND-L6-B2 DATED 31-01-2025',
    'H9' : 'INV-STEG-LC-2025-001',
    'H10' : '19-02-2025',
    'B11' : 'STEG INTERNATIONAL SERVICES',
    'G11' : '122-855-929',
    'G12' : '40-018812-P',
}
#从数据源获取的产品信息
items = {
    'item1' : {
        'SN': 1,
        'description' : 'LV ABC CABLE 4*95 sqmm',
        'unit':'KM',
        'qty': 78,
        'unit-price': 14295000.00
    },
    'item2' : {
        'SN': 2,
        'description' : 'LV ABC CABLE 4*50 sqmm',
        'unit':'KM',
        'qty': 30,
        'unit-price': 8074000.00
    },
    'item3': {
        'SN': 3,
        'description': 'LV ABC CABLE 4*25 sqmm',
        'unit': 'KM',
        'qty': 30,
        'unit-price': 123.00
    }
}

insertRowNum = 16
# 定义线的样式
thin = Side(border_style="thin", color="000000")  # 黑色细线
border = Border(left=thin, right=thin, top=thin, bottom=thin)

cells = [1,2,6,7,8]
img = Image(img_path)

img.width = 710
img.height = 146

ws.add_image(img,"A1")

def fill_other_cells(other_data):
    for key, value in other_data.items():
        ws[key].value = value

def insert_rows(row_num, row_details):
    all_row_cells = []
    for _ in range(len(items)):
        cells_in_row = []
        for letter in "ABCDEFGHI":
            col = ws[f"{letter}{row_num}"]
            col.border = border
            cells_in_row.append(col)

        all_row_cells.append(cells_in_row)
        row_num +=1
    return all_row_cells

def insert_values (origin_obj_list,to_be_filled_obj_list,to_be_merged_obj_list,item_dic):
    for row_cells, item in zip(to_be_filled_obj_list, item_dic.values()):
        values_to_insert = list(item.values())
        for cell, value in zip(row_cells, values_to_insert):
            cell.value = value
        for cell in origin_obj_list:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border



# for i, (key, subdict) in enumerate(items.items()):
    #     row = start_row + i
    #     # 只取子字典的 value，保证顺序和 cells 一致
    #     for col, value in zip(cells_list, subdict.values()):
    #         ws.cell(row=row, column=col, value=value)

fill_other_cells(data)
created_cells = insert_rows(insertRowNum, items)

c_to_e_list = []
i_list = []
for row in created_cells:
    # 提取 C~E
    extracted = row[2:5]
    extracted2 = row[-1]
    c_to_e_list.append(extracted)
    i_list.append(extracted2)

    # 从原行删除 C~E
    row[2:5] = []

insert_values(created_cells,items)
last_row = ws.max_row
ws.print_area = f"A1:I{last_row}"

wb.save(new_file_path)
