# 调用库
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side
from openpyxl.styles import Alignment

loading_path = 'xlsx/test1.xlsx'
saving_path = 'xlsx/test2.xlsx'
image_path = 'images/title.png'

wb = load_workbook(loading_path)
ws = wb['Sheet1']
#准备数据
#从数据源获取的其他信息
other_data={
    'C4' : '002LCNB250350001',
    'D5' : 'EC-202501-STEG-LC-COND-L6-B2 DATED 31-01-2025',
    'H9' : 'INV-STEG-LC-2025-001',
    'H10' : '19-02-2025',
    'B11' : 'STEG INTERNATIONAL SERVICES',
    'G11' : '122-855-929',
    'G12' : '40-018812-P',
}
#从数据源获取的产品信息
items = {
    'item1' : {
        'SN': 1,
        'description' : 'LV ABC CABLE 4*95 sqmm',
        'unit':'KM',
        'qty': 78,
        'unit-price': 14295000.00
    },
    'item2' : {
        'SN': 2,
        'description' : 'LV ABC CABLE 4*50 sqmm',
        'unit':'KM',
        'qty': 30,
        'unit-price': 8074000.00
    },
    'item3': {
        'SN': 3,
        'description': 'LV ABC CABLE 4*25 sqmm',
        'unit': 'KM',
        'qty': 30,
        'unit-price': 123.00
    }
}

starting_row = 16
# 配置
def configs() -> None:
    # 配置打印和样式
    ws.print_title_rows = "1:1"
    last_row = ws.max_row
    ws.print_area = f"A1:I{last_row}"
    
#1.写入其他数据
def input_data_others(other_data:dict,image_path:str) -> None: 
    img = Image(image_path)
    img.width = 710
    img.height = 146
    ws.add_image(img,"A1")
    for key, value in other_data.items():
        ws[key].value = value # type: ignore
        
#2. 根据len(details_data)确认需要插入的行数,并插入行
def insert_rows(starting_row:int,rows:dict) -> None:
    for i in range(len(rows)):
        ws.insert_rows(starting_row)
        
#3. 获取插入的所有单元格对象,return一个包含所有单元格对象的列表
def get_cells_obj(starting_row:int,rows:dict) -> list:
    all_cells_obj = []
    for _ in range(len(rows)):
        one_row_cells = []
        for letter in "ABCDEFGHI":
            row_cell = ws[f'{letter}{starting_row}']
            one_row_cells.append(row_cell)
        starting_row += 1
        all_cells_obj.append(one_row_cells)
    # print(len(all_cells_obj))
    # print(all_cells_obj)
    return all_cells_obj
            
            
#4. 将列表中单元格对象分类:需要填入details_data的列表,需要填入公式的列表
def slice_obj_list(original_obj_list:list):
    merge_list = [row[1:5] for row in original_obj_list]
    formular_list = [row[8] for row in original_obj_list]
    for row in original_obj_list:
        del row[1:5]
    data_list = original_obj_list
    print(merge_list)
    print('______________________________')
    print(formular_list)
    print('______________________________')
    print(data_list)
    
    
    # return merge_list, formular_list, data_list
    # print(merge_list)
    # print(original_obj_list)
    
#5. 处理detailed_data, 返回一个2元列表
def get_values_from_dic(details_data):
    ...
#6. 为单元格填入数据
def insert_value(cells_obj_list,value_list):
    ...
#7. 为单元格添加公式
def insert_formula(formula_obj_list,formula):
    ...
#8. 为单元格添加边框
def add_border(cells_obj_list):
    ...

def main():
    # input_data_others(other_data,image_path)
    # insert_rows(starting_row,items)
    all_obj_list = get_cells_obj(starting_row,items)
    slice_obj_list(all_obj_list)
    # configs()

if __name__ == '__main__':
    main()
    
# wb.save(saving_path)